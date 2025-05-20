from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Medicine, MedicineSale, Category, Supplier, UserProfile
from datetime import datetime
# from django.utils import timezone
from django.db.models import Sum, F, ExpressionWrapper, DecimalField

from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.utils import get_column_letter

from .forms import MedicineForm, ProfileForm

def get_base_context(request=None):
    now = datetime.now()
    context = {
        'day': now.strftime('%A'),
        'current_time': now.strftime('%I:%M %p'),
        'current_date': now.strftime('%B %d, %Y')
    }
    
    if request and request.user.is_authenticated:
        try:
            profile = request.user.profile
            context.update({
                'user_role': profile.role.get_name_display(),
                'user_position': profile.position,
                'user_department': profile.department
            })
        except UserProfile.DoesNotExist:
            context.update({
                'user_role': 'Guest',
                'user_position': '',
                'user_department': ''
            })
    
    return context


@login_required
def home(request):
    context = get_base_context(request)
    return render(request, 'base.html', context)


@login_required
def profile(request):
    context = get_base_context(request)
    profile = request.user.profile
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = ProfileForm(instance=profile)
    
    return render(request, 'layout/profile.html', context)


def user_login(request):
    context = get_base_context(request)
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                auth_login(request, user)
                messages.success(request, f'Welcome back, {username}!')
                return redirect('home')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()
    context['form'] = form
    return render(request, 'layout/login.html', context)

def user_logout(request):
    auth_logout(request)
    messages.info(request, 'You have been logged out successfully.')
    return redirect('user_login')

@login_required
def medicine_list(request):
    context = get_base_context(request)
  
    medicines = Medicine.objects.all()
    context.update({
        'medicines_count': Medicine.objects.count(),
    })
    if request.method == 'POST':
        search_query = request.POST.get('search')
        if search_query:
            medicines = medicines.filter(name__icontains=search_query)
            context['search_query'] = search_query
        else:
            messages.error(request, 'Please enter a valid search term.')
    context['medicines'] = medicines
    return render(request, 'layout/medicines.html', context)


@login_required
def add_medicine(request):
    context = get_base_context(request)

    if request.method == 'POST':
        form = MedicineForm(request.POST, request.FILES)
        if form.is_valid():
            medicine = form.save(commit=False)
            medicine.created_at = datetime.now()
            medicine.updated_at = datetime.now()
            medicine.save()
            supplier = form.cleaned_data.get('supplier')
            if supplier:
                supplier.save()
            messages.success(request, 'Medicine added successfully!')
            return redirect('medicines')
        else:
            messages.error(request, 'Error adding medicine. Please check the form.')
    else:
        form = MedicineForm()

    context['form'] = form  
    return render(request, 'layout/add_medicine.html', context)


@login_required
def edit_medicine(request, medicine_id):
    context = get_base_context(request)
    medicine = get_object_or_404(Medicine, id=medicine_id)
    if request.method == 'POST':
        form = MedicineForm(request.POST, instance=medicine)
        if form.is_valid():
            form.save()
            messages.success(request, 'Medicine updated successfully!')
            return redirect('medicines')
        else:
            messages.error(request, 'Error updating medicine. Please check try again!')
    else:
        form = MedicineForm(instance=medicine)
    context['form'] = form
   
    return render(request, 'layout/edit_medicine.html', context)


@login_required
def delete_medicine(request, medicine_id):
    context = get_base_context(request)
    medicine = get_object_or_404(Medicine, id=medicine_id)
    if request.method == 'POST':
        medicine.delete()
        messages.success(request, 'Medicine deleted successfully!')
        return redirect('medicines')
    context['medicine'] = medicine
   
    return render(request, 'pharmacy/delete_medicine.html', context)

@login_required
def medicine_details(request, medicine_id):
    context = get_base_context(request)
    medicine = Medicine.objects.get(id=medicine_id)
    context['medicine'] = medicine
    return render(request, 'layout/medicine_details.html', context)


@login_required
def dashboard(request):
    context = get_base_context(request)
    total_revenue = MedicineSale.objects.aggregate(
        total=Sum(
            ExpressionWrapper(
                F('quantity_sold') * F('price_per_unit'),
                output_field=DecimalField()
            )
        )
    )['total'] or 0
    
    context.update({
        'medicines_count': Medicine.objects.count(),
        'categories_count': Category.objects.count(),
        'suppliers_count': Supplier.objects.count(),
        'recent_medicines': Medicine.objects.order_by('-created_at')[:5],
        # 'users_count': UserProfile.objects.count(),
    })
    return render(request, 'layout/dashboard.html', context)


# PDF View
def download_pdf_report(request):
    medicines = Medicine.objects.all()
    template_path = 'layout/reports/pdf_template.html'  # You’ll create this
    context = {'medicines': medicines}
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="medicine_report.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


# Excel View
def download_excel_report(request):
    medicines = Medicine.objects.all()

    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medicine Report"

    # Define headers
    headers = [
        'Name',
        'Batch Number',
        'Manufacturer',
        'Quantity',
        'Price',
        'Manufactured Date',
        'Expiry Date',
        'Created At',
    ]
    ws.append(headers)

    # Fill worksheet with data
    for med in medicines:
        ws.append([
            med.name,
            med.batch_number,
            med.manufacturer,
            med.quantity,
            float(med.price) if med.price else '',
            med.manufactured_date.strftime("%Y-%m-%d") if med.manufactured_date else '',
            med.expiry_date.strftime("%Y-%m-%d") if med.expiry_date else '',
            med.created_at.strftime("%Y-%m-%d %H:%M") if med.created_at else '',
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=medicine_report.xlsx'
    wb.save(response)
    return response